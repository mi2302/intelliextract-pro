import zipfile
import os
import sys
import json
import re
import shutil
import tempfile
from typing import List, Dict, Any
from openpyxl import load_workbook

# MONKEY PATCH: Oracle's VML files are often invalid XML (e.g. <br> instead of <br/>).
# We force lxml to use recovery mode so openpyxl doesn't crash during load/save.
try:
    from lxml import etree
    _original_fromstring = etree.fromstring
    def _safe_fromstring(text, parser=None):
        if parser is None:
            if isinstance(text, str): text = text.encode('utf-8')
            parser = etree.XMLParser(recover=True, remove_comments=False)
        return _original_fromstring(text, parser=parser)
    etree.fromstring = _safe_fromstring
    print("[PYTHON] LXML Recovery Mode enabled (Surgical).")
except ImportError:
    print("[PYTHON] Warning: lxml not found. XML syntax errors in templates may cause crashes.", file=sys.stderr)

def surgical_populate(template_path: str, output_path: str, specs: List[Dict[str, Any]]):
    temp_dir = tempfile.mkdtemp()
    
    def sanitize_value(v):
        if v is None: return ""
        if isinstance(v, str):
            v = "".join(ch for ch in v if ch.isprintable() or ch in "\n\r\t")
            v = v.replace('\u0000', '').replace('\uFFFE', '').replace('\uFFFF', '')
        return v

    try:
        # Step 1: Detect Style
        data_cell_style = None
        try:
            with zipfile.ZipFile(template_path, 'r') as zf:
                # Find any worksheet that might have a style on row 5
                for name in zf.namelist():
                    if name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
                        content = zf.read(name).decode('utf-8')
                        style_match = re.search(r'<c [^>]*r="[A-Z]5"[^>]*s="([^"]+)"', content)
                        if not style_match:
                            style_match = re.search(r'<c [^>]*s="([^"]+)"', content)
                        if style_match:
                            data_cell_style = style_match.group(1)
                            print(f"[PYTHON] Detected template style: s=\"{data_cell_style}\"")
                            break
        except: pass

        # Step 2: Generate Data using OpenPyXL
        temp_populated = os.path.join(temp_dir, "populated.xlsm")
        wb = load_workbook(template_path, keep_vba=False)
        
        populated_any = False
        for spec in specs:
            sheet_name = spec.get('sheetName')
            if sheet_name not in wb.sheetnames:
                print(f"[PYTHON] Warning: Sheet '{sheet_name}' not found.")
                continue
            ws = wb[sheet_name]
            data = spec.get('data', [])
            columns = spec.get('columns', [])
            if not data: continue
            
            populated_any = True
            for row_idx, row_record in enumerate(data, start=5):
                for col_idx, col_spec in enumerate(columns, start=1):
                    val = row_record.get(col_spec.get('targetName', col_spec.get('alias')))
                    val = sanitize_value(val)
                    if val is not None:
                        ws.cell(row=row_idx, column=col_idx, value=val)
        
        if not populated_any:
            print("[PYTHON] No data records found to populate.", file=sys.stderr)
            return False

        wb.save(temp_populated)
        wb.close()
        
        # Step 3: Precise Injection
        modified_xmls = {}
        with zipfile.ZipFile(temp_populated, 'r') as zext:
            with zipfile.ZipFile(template_path, 'r') as zf_orig:
                # 3.1: Map Sheet Names to Paths (OpenPyXL version)
                z_wb_xml = zext.read('xl/workbook.xml').decode('utf-8')
                z_rels_xml = zext.read('xl/_rels/workbook.xml.rels').decode('utf-8')
                z_rel_map = {m.group(1): m.group(2) for m in re.finditer(r'Id=["\']([^"\']+)["\'][^>]*Target=["\']([^"\']+)["\']', z_rels_xml)}
                z_name_to_path = {}
                for m in re.finditer(r'<sheet [^>]*name=["\']([^"\']+)["\'][^>]*r:id=["\']([^"\']+)["\']', z_wb_xml):
                    target = z_rel_map.get(m.group(2))
                    if target: z_name_to_path[m.group(1)] = f"xl/{target}" if target.startswith('worksheets/') else f"xl/worksheets/{target}"

                # 3.2: Map Sheet Names to Paths (Original version)
                o_wb_xml = zf_orig.read('xl/workbook.xml').decode('utf-8')
                o_rels_xml = zf_orig.read('xl/_rels/workbook.xml.rels').decode('utf-8')
                o_rel_map = {m.group(1): m.group(2) for m in re.finditer(r'Id=["\']([^"\']+)["\'][^>]*Target=["\']([^"\']+)["\']', o_rels_xml)}
                o_name_to_path = {}
                for m in re.finditer(r'<sheet [^>]*name=["\']([^"\']+)["\'][^>]*r:id=["\']([^"\']+)["\']', o_wb_xml):
                    target = o_rel_map.get(m.group(2))
                    if target: o_name_to_path[m.group(1)] = f"xl/{target}" if target.startswith('worksheets/') else f"xl/worksheets/{target}"

                # 3.3: Sync Content Types
                try:
                    opyxl_ct = zext.read('[Content_Types].xml').decode('utf-8')
                    orig_ct = zf_orig.read('[Content_Types].xml').decode('utf-8')
                    if 'sharedStrings.xml' in opyxl_ct and 'sharedStrings.xml' not in orig_ct:
                        entry = '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
                        modified_xmls['[Content_Types].xml'] = orig_ct.replace('</Types>', f'  {entry}\n</Types>')
                except: pass

                # 3.4: Inject sheetData
                processed_sheets = 0
                for s_name, z_path in z_name_to_path.items():
                    o_path = o_name_to_path.get(s_name)
                    if not o_path: continue
                    
                    try:
                        z_cont = zext.read(z_path).decode('utf-8')
                        o_cont = zf_orig.read(o_path).decode('utf-8')
                        
                        # Support both <sheetData>...</sheetData> and self-closing <sheetData/>
                        s_data_match = re.search(r'<sheetData.*?(/>|</sheetData>)', z_cont, re.DOTALL)
                        s_dim_match = re.search(r'<dimension ref=["\'][^"\']+["\']/>', z_cont)
                        
                        if s_data_match and s_dim_match:
                            s_data = s_data_match.group(0)
                            s_dim = s_dim_match.group(0)
                            if data_cell_style:
                                s_data = re.sub(r'(<c [^>]*r="[A-Z][0-9]{1,6}"(?![^>]*s="))', r'\1 s="' + data_cell_style + '"', s_data)
                            
                            new_c = re.sub(r'<dimension ref=["\'][^"\']+["\']/>', s_dim, o_cont)
                            new_c = re.sub(r'<sheetData.*?(/>|</sheetData>)', s_data, new_c, flags=re.DOTALL)
                            modified_xmls[o_path] = new_c
                            processed_sheets += 1
                    except Exception as e:
                        print(f"[PYTHON] Error mapping sheet {s_name}: {e}")

                if processed_sheets == 0:
                    print("[PYTHON] Warning: Matched zero sheets for injection.", file=sys.stderr)

                if 'xl/sharedStrings.xml' in zext.namelist():
                    modified_xmls['xl/sharedStrings.xml'] = zext.read('xl/sharedStrings.xml').decode('utf-8')

        # Step 4: Final Rebuild
        temp_final = os.path.join(temp_dir, "final.xlsm")
        with zipfile.ZipFile(template_path, 'r') as zin:
            with zipfile.ZipFile(temp_final, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename in modified_xmls:
                        zout.writestr(item, modified_xmls.pop(item.filename).encode('utf-8'))
                    else:
                        zout.writestr(item, zin.read(item.filename))
                for f, c in modified_xmls.items():
                    zout.writestr(f, c.encode('utf-8'))
        
        if os.path.exists(output_path): os.remove(output_path)
        shutil.move(temp_final, output_path)
        print(f"Success: Template Populated to {output_path}")
        return True

    except Exception as e:
        print(f"Population Error: {str(e)}", file=sys.stderr)
        import traceback; traceback.print_exc()
        return False
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        sys.exit(1)
    t_path, o_path, s_input = sys.argv[1:4]
    try:
        if os.path.exists(s_input):
            with open(s_input, 'r', encoding='utf-8') as f: specs_data = json.load(f)
        else: specs_data = json.loads(s_input)
    except: sys.exit(1)
    if not surgical_populate(t_path, o_path, specs_data): sys.exit(1)
