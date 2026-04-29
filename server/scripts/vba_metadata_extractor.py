import sys
import json
import os
import re

def apply_openpyxl_fix():
    # CRITICAL: Apply openpyxl fix BEFORE importing openpyxl
    try:
        from lxml import etree
        import openpyxl.xml
        import openpyxl.writer.excel
        import openpyxl.reader.excel
        
        def safe_fromstring(text):
            parser = etree.XMLParser(recover=True, encoding='utf-8')
            if isinstance(text, bytes):
                text_str = text.decode('utf-8', errors='ignore')
            else:
                text_str = text
            
            # Fix common Oracle FBDI XML issues: corrupted tags, namespaces, etc.
            text_str = text_str.strip()
            text_str = re.sub(r'<(br|font|/font)[^>]*>', ' ', text_str, flags=re.IGNORECASE)
            text_str = re.sub(r'<(v:(?:textbox|shape|imagedata))[^>]*>.*?</\1>', r'<\1/>', text_str, flags=re.DOTALL | re.IGNORECASE)
            text_str = re.sub(r' xmlns(:.*?)?=""', '', text_str) # Remove empty namespaces
            
            try:
                return etree.fromstring(text_str.encode('utf-8'), parser=parser)
            except Exception:
                # Absolute fallback for stylesheet/styles.xml: return a dummy valid styles root
                if 'styleSheet' in text_str:
                    return etree.fromstring(b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"></styleSheet>')
                return etree.fromstring(b'<{http://schemas.openxmlformats.org/spreadsheetml/2006/main}xml/>')

        openpyxl.xml.fromstring = safe_fromstring
        openpyxl.writer.excel.fromstring = safe_fromstring
        openpyxl.xml.LXML = True
    except Exception:
        pass

import openpyxl  # type: ignore
from oletools.olevba import VBA_Parser  # type: ignore

def idx_to_letter(n):
    """Convert a 1-based column index to an Excel column letter."""
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def extract_vba_constants(target_path):
    try:
        vbaparser = VBA_Parser(target_path)
        macros = list(vbaparser.extract_macros()) if hasattr(vbaparser, 'extract_macros') else []
        
        results = {
            "startingDataRowNumber": 4, # Default
            "zipFileName": "FBDI_Import",
            "dataOnlySheets": [],
            "columnShifts": []
        }

        for (filename, stream_path, vba_filename, vba_code) in macros:
            # zipFileName - Pattern A (Constants)
            match = re.search(r'Public\s+Const\s+zipFileName\s+As\s+String\s*=\s*"(.*?)"', vba_code, re.IGNORECASE)
            if match:
                results['zipFileName'] = match.group(1)
            
            # zipFileName - Pattern B (GetSaveAsFilename)
            if results['zipFileName'] == "FBDI_Import":
                match = re.search(r'GetSaveAsFilename\("([^"]+)"', vba_code, re.IGNORECASE)
                if match:
                    results['zipFileName'] = match.group(1).strip()
            
            # startingDataRowNumber
            match = re.search(r'Public Const startingDataRowNumber As Long = (\d+)', vba_code, re.IGNORECASE)
            if match:
                results['startingDataRowNumber'] = int(match.group(1))

            # dataOnlySheets - Multiple Patterns
            patterns = [
                r'dataOnlySheets\((\d+)\)\s*=\s*"(.*?)"',
                r'Counter\s*=\s*(\d+)\s+Then.*?\.Name\s*=\s*"(.*?)"',
                r'Sheets\(SHCOUNT\s*\+\s*(\d+)\)\.Name\s*=\s*"(.*?)"'
            ]
            for p in patterns:
                matches = re.findall(p, vba_code, re.DOTALL | re.IGNORECASE)
                if matches:
                    sorted_m = sorted(matches, key=lambda x: int(x[0]))
                    for idx, name in sorted_m:
                        if name not in results['dataOnlySheets']:
                            results['dataOnlySheets'].append(name)

            # --- Improved Column Shift Detection ---
            # Find all GenCSV loops (Sales Order has multiple)
            loops = re.findall(r'For\s+Counter\s*=\s*1\s+To\s+SHCOUNT\s*-\s*1(.*?)(?:Next\s+Counter|Next)', vba_code, re.DOTALL | re.IGNORECASE)
            
            for loop_content in loops:
                # Split loop content into segments by 'If Counter = N'
                segments = re.split(r'If\s+Counter\s*=\s*(\d+)\s+Then', loop_content, flags=re.IGNORECASE)
                
                for i in range(len(segments)):
                    if i == 0:
                        segment_code = segments[i]
                        ctx = "All"
                    elif i % 2 == 1:
                        continue # Sheet number
                    else:
                        segment_code = segments[i]
                        ctx = segments[i-1]
                    
                    extract_shifts_from_block(segment_code, ctx, results)
            
            # Global shifts fallback
            extract_shifts_from_block(vba_code, "Global", results)
        
        # Deduplicate
        # --- Final Deduplication and Cleanup ---
        unique_shifts = []
        seen_keys = {} # logic_key -> index in unique_shifts
        
        for s in results['columnShifts']:
            # Create a "logic key" (type, source/start, target/end)
            l_key = (s.get('type'), s.get('sourceCol') or s.get('startCol'), s.get('targetCell') or s.get('endCol'))
            target = str(s.get('TargetSheet', ''))

            if l_key not in seen_keys:
                seen_keys[l_key] = len(unique_shifts)
                unique_shifts.append(s)
            else:
                idx = seen_keys[l_key]
                existing_target = str(unique_shifts[idx].get('TargetSheet', ''))
                # Prefer All/Specific over Global
                if existing_target.lower() == "global" and target.lower() != "global":
                    unique_shifts[idx] = s

        # Remove redundant DELETE_COL if it's the source of a MOVE_COL
        move_srcs = {s['sourceCol'] for s in unique_shifts if s['type'] == 'MOVE_COL' and s.get('deleteOriginal')}
        results['columnShifts'] = [s for s in unique_shifts if not (s['type'] == 'DELETE_COL' and s['startCol'] in move_srcs and s['startCol'] == s['endCol'])]

        vbaparser.close()
        return results
    except Exception as e:
        print(f"VBA Parser error: {e}", file=sys.stderr)
        return {}

def extract_shifts_from_block(block_code, ctx, results):
    # Normalize block code: remove underscores for multiline
    code = block_code.replace(' _\n', ' ').replace(' _\r\n', ' ')

    # 1. MOVE_COL Detection (Improved)
    # Pattern A: Select -> Copy/Cut -> Select -> Insert
    move_pattern = (
        r'(?:Columns\("?([A-Z]+):?[A-Z]*"?\)|Range\("?([A-Z]+)\d+"?\)|Range\(Cells\(\d+,\s*(\d+|[\w\d]+)\),\s*Cells\([\w\d]+\s*,\s*(\d+|[\w\d]+)\)\))\.Select'
        r'.*?'
        r'Selection\.(?:Copy|Cut)'
        r'.*?'
        r'(?:Range\("?([A-Z]+)\d+"?\)|Cells\(\d+,\s*(\d+|[\w\d]+)\))\.Select'
        r'\s*Selection\.Insert Shift:=xlToRight'
    )
    for m in re.finditer(move_pattern, code, re.DOTALL | re.IGNORECASE):
        src_raw = m.group(1) or m.group(2) or m.group(3) or m.group(4)
        tgt_raw = m.group(5) or m.group(6)
        src = src_raw if not src_raw.isdigit() else idx_to_letter(int(src_raw))
        tgt = tgt_raw if not tgt_raw.isdigit() else idx_to_letter(int(tgt_raw))
        results['columnShifts'].append({
            "type": "MOVE_COL", "sourceCol": src, "targetCell": tgt,
            "deleteOriginal": "Selection.Delete" in code[m.end():m.end()+200] or ".Cut" in m.group(0),
            "TargetSheet": ctx
        })

    # 2. DELETE_COL Detection (Improved to handle direct calls)
    # Pattern A: Columns("A:A").Delete or Range("A:B").Delete
    for m in re.finditer(r'(?:Columns\("([A-Z]+):([A-Z]+)"\)|Range\("([A-Z]+):([A-Z]+)"\))\.Delete', code, re.IGNORECASE):
        s1, e1, s2, e2 = m.groups()
        start, end = s1 or s2, e1 or e2
        results['columnShifts'].append({"type": "DELETE_COL", "startCol": start, "endCol": end or start, "TargetSheet": ctx})

    # Pattern B: Columns/Range/Cells.Select -> Selection.Delete
    del_sel_p = r'(?:Columns\("([A-Z]+):([A-Z]+)"\)|Range\("([A-Z]+):([A-Z]+)"\)|Range\(Cells\(1,\s*(\d+|[\w\d]+)\),\s*Cells\([\w\d]+,\s*(\d+|[\w\d]+)\)\))\.Select\s*Selection\.Delete'
    for m in re.finditer(del_sel_p, code, re.IGNORECASE):
        s1, e1, s2, e2, si, ei = m.groups()
        start = s1 or s2 or (idx_to_letter(int(si)) if si and si.isdigit() else si)
        end = e1 or e2 or (idx_to_letter(int(ei)) if ei and ei.isdigit() else ei)
        if start:
            results['columnShifts'].append({"type": "DELETE_COL", "startCol": start, "endCol": end or start, "TargetSheet": ctx})

    # 3. HIDE_COL Detection
    hide_pattern = r'(?:Columns\("([^"]+)"\)|Range\("([^"]+)"\)\.EntireColumn)\.Hidden\s*=\s*(True|False|Not|[\w\d]+)'
    for m in re.finditer(hide_pattern, code, re.IGNORECASE):
        cols, rng, state = m.groups()
        if state.lower() in ['true', 'not'] or not state[0].isdigit():
            final = cols or rng
            for part in final.split(','):
                if ':' in part:
                    s, e = part.strip().split(':')
                    results['columnShifts'].append({"type": "HIDE_COL", "startCol": s, "endCol": e, "TargetSheet": ctx})
                else:
                    p = part.strip()
                    results['columnShifts'].append({"type": "HIDE_COL", "startCol": p, "endCol": p, "TargetSheet": ctx})

    # 4. HIDE_ROW Detection
    hide_row_pattern = r'Rows\("(\d+):(\d+)"\)\.Hidden\s*=\s*(True|False)'
    for match in re.finditer(hide_row_pattern, block_code, re.IGNORECASE):
        start_row, end_row, state = match.groups()
        if state.lower() == 'true':
            results['columnShifts'].append({"type": "HIDE_ROW", "startRow": int(start_row), "endRow": int(end_row), "TargetSheet": ctx})

    # 5. FORMAT_COL Detection
    format_pattern = r'Columns\("([A-Z]+):([A-Z]+)"\)\.NumberFormat\s*=\s*"(.*?)"'
    for match in re.finditer(format_pattern, block_code, re.IGNORECASE):
        start_col, end_col, fmt = match.groups()
        results['columnShifts'].append({"type": "FORMAT_COL", "startCol": start_col, "endCol": end_col, "format": fmt, "TargetSheet": ctx})

def extract_excel_structure(target_path, starting_row=4):
    apply_openpyxl_fix()
    try:
        wb = openpyxl.load_workbook(target_path, data_only=True, keep_vba=False)
        sheets_info = {}
        
        for sheet_name in wb.sheetnames:
            if any(x in sheet_name.lower() for x in ['instruction', 'reference', 'label', 'note']):
                continue
            
            ws = wb[sheet_name]
            headers = []
            # Extract headers from the specified starting_row (usually 4)
            for c_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=starting_row, column=c_idx).value
                headers.append(str(val).strip() if val is not None else "")
            
            while headers and not headers[-1]:
                headers.pop()

            if any(headers):
                sheets_info[sheet_name] = {
                    "headers": headers,
                    "columnCount": len(headers)
                }
        
        return sheets_info
    except Exception as e:
        print(f"Excel error: {e}", file=sys.stderr)
        return {}

def main():
    if len(sys.argv) < 2:
        sys.exit(1)
        
    target_path = sys.argv[1]
    vba_meta = extract_vba_constants(target_path)
    # Use the extracted starting row for excel parsing
    excel_meta = extract_excel_structure(target_path, vba_meta.get("startingDataRowNumber", 4))
    
    final_metadata = {
        "zipFileName": vba_meta.get("zipFileName"),
        "startingDataRowNumber": vba_meta.get("startingDataRowNumber"),
        "dataOnlySheets": vba_meta.get("dataOnlySheets"),
        "columnShifts": vba_meta.get("columnShifts"),
        "sheets": excel_meta
    }
    
    print(json.dumps(final_metadata, indent=2))

if __name__ == "__main__":
    main()
