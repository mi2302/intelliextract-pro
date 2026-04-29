import sys
import json
import os
import openpyxl  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from typing import List, Dict, Any, Union

def populate_fbdi_internal(template_path: str, output_path: str, specs: List[Dict[str, Any]]) -> bool:
    """
    Internal population logic that supports a macro-free fallback.
    """
    try:
        # Load the existing template
        # We hardcode keep_vba=True for .xlsm files to ensure 100% macro preservation.
        is_xlsm = template_path.lower().endswith('.xlsm')
        keep_vba = True if is_xlsm else False
        
        # WORKAROUND: Force openpyxl to use a lenient LXML parser for VML.
        # This prevents "tag mismatch" errors by enabling LXML's recovery mode,
        # ensuring we keep the Oracle buttons/logos without mixed-parser TypeErrors.
        try:
            from lxml import etree
            import openpyxl.xml
            import openpyxl.writer.excel
            import re

            # Create a lenient parser that ignores syntax errors (like unclosed <br> tags)
            lenient_parser = etree.XMLParser(recover=True, no_network=True)
            
            def safe_fromstring(text):
                try:
                    if isinstance(text, str):
                        text = text.encode('utf-8')
                    return etree.fromstring(text, parser=lenient_parser)
                except Exception:
                    # Fallback to regex repair if even recovery fails
                    text_str = text.decode('utf-8', errors='ignore') if isinstance(text, bytes) else text
                    text_str = re.sub(r'<(br|font|/font)[^>]*>', ' ', text_str, flags=re.IGNORECASE)
                    return etree.fromstring(text_str.encode('utf-8'), parser=lenient_parser)

            # Patch openpyxl to use our safe/lenient reader
            openpyxl.xml.fromstring = safe_fromstring
            openpyxl.writer.excel.fromstring = safe_fromstring
            
            # Ensure openpyxl knows it's using LXML so it doesn't mix with ElementTree
            openpyxl.xml.LXML = True 
            print("[PYTHON] Applied lenient LXML parser for FBDI compatibility.")
        except Exception as patch_err:
            print(f"[PYTHON] Lenient LXML patch failed: {str(patch_err)}", file=sys.stderr)

        # openpyxl will use lxml if installed, which is more robust
        wb = openpyxl.load_workbook(template_path, keep_vba=keep_vba)
        
        for spec in specs:
            sheet_name: str = str(spec.get('sheetName', ''))
            data: List[Dict[str, Any]] = spec.get('data', [])
            
            if not sheet_name or sheet_name not in wb.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not found in workbook. Skipping.", file=sys.stderr)
                continue
                
            ws = wb[sheet_name]
            
            # Find the header row (scan first 20 rows)
            header_row_idx: int = -1
            headers: List[str] = []
            
            cols_metadata: List[Dict[str, Any]] = spec.get('columns', [])
            target_name_match: str = ''
            if cols_metadata and len(cols_metadata) > 0:
                target_name_match = str(cols_metadata[0].get('targetName', '')).replace('*', '')
            
            for r_idx in range(1, 21):
                row_cells = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, ws.max_column + 1)]
                valid_cells = [str(c).strip() for c in row_cells if c is not None and str(c).strip()]
                
                if target_name_match and any(target_name_match in str(c) for c in valid_cells if c):
                    header_row_idx = r_idx
                    headers = [str(c).strip() if c else '' for c in row_cells]
                    break
            
            if header_row_idx == -1:
                header_row_idx = 1 # Fallback
                headers = [str(ws.cell(row=1, column=i).value).strip() if ws.cell(row=1, column=i).value else '' for i in range(1, ws.max_column + 1)]
            
            # Map column indices for faster lookup
            header_map: Dict[str, int] = {name: idx+1 for idx, name in enumerate(headers) if name}
            current_write_row: int = header_row_idx + 1
            
            for row_record in data:
                # Type safe iteration over items
                items_to_write: Dict[str, Any] = row_record
                for col_name, val in items_to_write.items():
                    target_col: Union[int, None] = header_map.get(col_name)
                    if not target_col:
                        # Fuzzy match (with/without *)
                        clean_col_name = col_name.replace('*', '')
                        for h_name, h_idx in header_map.items():
                            if h_name.replace('*', '') == clean_col_name:
                                target_col = h_idx
                                break
                    
                    if target_col is not None:
                        ws.cell(row=current_write_row, column=target_col).value = val
                
                # Explicit increment with type ignore to satisfy quirky analyzer
                current_write_row: int = current_write_row + 1 # type: ignore
            
            # Self-healing parser (safe_fromstring) handles the VML malformed tags
            # so we no longer need to strip comments. This restores instruction bubbles.
            pass
                
        # Save the populated workbook
        # We always use the requested output_path to ensure Node.js finds it.
        # If we dropped VBA, the file will still be saved to output_path (even if it has .xlsm ext)
        # for standard Excel/openpyxl this is valid, it just won't have macros.
        wb.save(output_path)
        print(f"[PYTHON] Data population complete. Starting Hybrid Metadata Restoration...")

        # HYBRID RESTORATION: OpenPyXL often corrupts VML/Comments in .xlsm files.
        # We now surgically restore the ORIGINAL VML and Comment files from the template
        # into the newly saved workbook. This ensures 100% preservation of buttons and logos.
        try:
            import zipfile
            import tempfile
            import shutil

            temp_dir = tempfile.mkdtemp()
            restored_zip = os.path.join(temp_dir, "restored.zip")
            
            # Map of files to restore bit-for-bit from the original template
            to_restore = [
                'xl/vbaProject.bin', # Always restore the macro project binary
                'xl/_rels/vbaProject.bin.rels'
            ]
            
            with zipfile.ZipFile(template_path, 'r') as zin:
                # Dynamically find all drawings and comments in the original
                for name in zin.namelist():
                    if 'comments' in name.lower() or 'drawings/vml' in name.lower() or 'drawings/drawing' in name.lower():
                        to_restore.append(name)
                    if name.endswith('.rels') and ('drawings' in name.lower() or 'worksheets' in name.lower()):
                        to_restore.append(name)

                with zipfile.ZipFile(output_path, 'r') as zin_new:
                    with zipfile.ZipFile(restored_zip, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                        # 1. Copy everything from the NEW file (which has the data)
                        for item in zin_new.infolist():
                            if item.filename not in to_restore:
                                zout.writestr(item, zin_new.read(item.filename))
                        
                        # 2. Overwrite with ORIGINAL metadata files
                        for item_name in set(to_restore):
                            try:
                                zout.writestr(item_name, zin.read(item_name))
                            except KeyError:
                                pass # File might not exist in template
            
            # Replace the OPyXL output with our surgically restored version
            shutil.move(restored_zip, output_path)
            shutil.rmtree(temp_dir)
            print(f"Success: Hybrid restoration complete. File saved to {output_path}")
            return True
        except Exception as restore_err:
            print(f"[PYTHON] Hybrid Restoration failed: {str(restore_err)}", file=sys.stderr)
            # Fallback: keep the OPyXL version (at least it has data)
            return True

    except Exception as e:
        # We no longer fall back silently. We want the user to have macros or know it failed.
        print(f"Population Error: {str(e)}", file=sys.stderr)
        raise e

def populate_fbdi(template_path: str, output_path: str, specs_input: str) -> bool:
    try:
        # specs_input can be raw JSON or a path to a JSON file
        try:
            if os.path.exists(specs_input):
                with open(specs_input, 'r', encoding='utf-8') as f:
                    specs: List[Dict[str, Any]] = json.load(f)
            else:
                specs: List[Dict[str, Any]] = json.loads(specs_input)
        except Exception as json_err:
            print(f"Error parsing specs: {str(json_err)}", file=sys.stderr)
            sys.exit(1)
            
        return populate_fbdi_internal(template_path, output_path, specs)
        
    except Exception as e:
        print(f"Critical error in Python script: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python populate_fbdi.py <template_path> <output_path> <specs_input>", file=sys.stderr)
        sys.exit(1)
        
    t_path: str = sys.argv[1]
    o_path: str = sys.argv[2]
    s_input: str = sys.argv[3]
    
    populate_fbdi(t_path, o_path, s_input)
